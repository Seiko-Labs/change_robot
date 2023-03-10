import openpyxl

# ! Credentials
from transliterate import translit

internet_hierarchy = ['PALO_ALTO_TOP2', 'INET_WITH_SOCIAL', 'PALO_ALTO_WITHOUT_SOCIAL', 'INET_STANDART',
                      'INET_JUST_HCSBK', 'INET_MIN_ACCESS']

xl = openpyxl.load_workbook('credentials.xlsx')
sheet = xl.active
login = sheet.cell(2, 1).value
password = sheet.cell(2, 2).value
colvir_adm_login = sheet.cell(6, 1).value
colvir_adm_password = sheet.cell(6, 2).value
colvir_login = sheet.cell(10, 1).value
colvir_password = sheet.cell(10, 2).value

# ! Filial maps
alph_map = {
    'ә': 'а', 'ғ': 'г', 'қ': 'q', 'ң': 'н', 'ө': 'о', 'ұ': 'у', 'ү': 'у', 'h': 'х', 'і': 'и', 'й': 'i',
    'Ә': 'А', 'Ғ': 'Г', 'Қ': 'Q', 'Ң': 'Н', 'Ө': 'О', 'Ұ': 'У', 'Ү': 'У', 'H': 'Х', 'І': 'И', 'Й': 'I',
}

bpm_exception = {
    'Видеоконсультант-эксперт 2': 'Видеоконсультант эксперт-2',
    'Работник УРБ/ОРБ. Исполнительные документы роль - (Работник УРБ/ОРБ. Исполнительные документы)': 'Работник УРБ/ОРБ. Исполнительные документы роль',
    'Пользователь HR': 'Участник',
    'Кредитный эксперт (роль - Кредитный эксперт)': 'Кредитный Эксперт',
    'Логисты (роль - Логист)': 'Логист'
}

colvir_map = {
    '01': 'Акмолинский ОФ АО «Жилищный строительный сберегательный банк «Отбасы банк»',
    '02': 'Актюбинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '03': 'Алматинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
    '04': 'Атырауский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
    '05': 'Восточно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '06': 'Жамбылский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '07': 'Западно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк',
    '08': 'Карагандинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '09': 'Костанайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '11': 'Кызылординский филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '12': 'Мангистауский ОФ АО «Жилищный строительный сберегательный банк «Отбасы Банк»',
    '13': 'Павлодарский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '14': 'Северо-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '15': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Алматы',
    '17': 'Абайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '18': 'АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '19': 'Центральный филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    '20': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Шымкент',
    '21': 'Туркестанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
}

bpm_map = {
    'Центральный аппарат': 'АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Мангистауский ОФ': 'Мангистауский ОФ АО «Жилищный строительный сберегательный банк «Отбасы Банк»',
    'Актюбинский ОФ': 'Актюбинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Филиал в г. Алматы': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Алматы',
    'Филиал в г. Астана': 'Центральный филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Атырауский ОФ': 'Атырауский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
    'Карагандинский ОФ': 'Карагандинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Акмолинский ОФ': 'Акмолинский ОФ АО «Жилищный строительный сберегательный банк «Отбасы банк»',
    'Костанайский ОФ': 'Костанайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Кызылординский ОФ': 'Кызылординский филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Павлодарский ОФ': 'Павлодарский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Северо-Казахстанский ОФ': 'Северо-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Абайский ОФ': 'Абайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Филиал в г. Шымкент': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Шымкент',
    'Алматинский ОФ': 'Алматинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
    'Жамбылский ОФ': 'Жамбылский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Туркестанский ОФ': 'Туркестанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Западно-Казахстанский ОФ': 'Западно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Восточно-Казахстанский ОФ': 'Восточно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"'
}

ad_map = {
    'CA': 'АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Aktau': 'Мангистауский ОФ АО «Жилищный строительный сберегательный банк «Отбасы Банк»',
    'Aktobe': 'Актюбинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Almaty': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Алматы',
    'Astana': 'Центральный филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Atyrau': 'Атырауский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
    'Karaganda': 'Карагандинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Kokshetau': 'Акмолинский ОФ АО «Жилищный строительный сберегательный банк «Отбасы банк»',
    'Kostanay': 'Костанайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Kyzylorda': 'Кызылординский филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Pavlodar': 'Павлодарский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Petropavlovsk': 'Северо-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Semey': 'Абайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Shymkent': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Шымкент',
    'Taldykorgan': 'Алматинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
    'Taraz': 'Жамбылский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Turkestan': 'Туркестанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Uralsk': 'Западно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Ust-Kamenogorsk': 'Восточно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"'
}

doc_map = {
    'Актау': 'Мангистауский ОФ АО «Жилищный строительный сберегательный банк «Отбасы Банк»',
    'Актобе': 'Актюбинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Алматы': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Алматы',
    'Астана': 'Центральный филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Атырау': 'Атырауский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
    'Караганда': 'Карагандинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Кокшетау': 'Акмолинский ОФ АО «Жилищный строительный сберегательный банк «Отбасы банк»',
    'Костанай': 'Костанайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Кызылорда': 'Кызылординский филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Павлодар': 'Павлодарский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Петропавловск': 'Северо-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Семей': 'Абайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Шымкент': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Шымкент',
    'Талдыкорган': 'Алматинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
    'Тараз': 'Жамбылский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Туркестан': 'Туркестанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Уральск': 'Западно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
    'Усть-Каменогорск': 'Восточно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"'
}

department_exceptions = [
    "Центр обслуживания №1",
    "Караганда Управление бэк-офиса",
    "Делопроизводитель ТОФ",
    "Отдел по работе с клиентами (универсал)",
    "Отдел развития бизнеса",
    "Специалист 1 категории - логист",
    "Специалист 1 категории информационных технологий",
    "Специалисты вне подразделения филиала ТОФ",
    "Центр обслуживания №1 филиала г. Экибастуз",
    "Центр обслуживания №1 г. Каскелен",
    "Центр обслуживания №2 Отеген Батыра",
    "ЦО Отеген Батыра",
    "ЦО №1 Акмолинского ОФ с.Косшы",
    "Специалисты по возмездному оказанию услуг",
    "Специалист по безопасности"
]


def get_group(data):
    wb = openpyxl.load_workbook('workgroups.xlsx')
    ws = wb.active
    for i in range(2, ws.max_row):
        if ws[f'A{i}'].value is not None and ws[f'A{i}'].value in data['department']:
            wb.close()
            return ws[f'B{i}'].value
    if data['filial_type'] == 'Пользователи ЦА':
        return '!!!!АРМ Смена пароля'
    elif data['filial_type'] == 'Пользователи филиалов':
        return '!!!АРМ Персонал (филиалы)'


def replace_filial(data):
    for i in doc_map.keys():
        if data['filial'] == doc_map[i]:
            data['filial_doc'] = i
    for i in ad_map.keys():
        if data['filial'] == ad_map[i]:
            data['filial_ad'] = i
    for i in colvir_map.keys():
        if data['filial'] == colvir_map[i]:
            data['filial_colvir'] = i
    for i in bpm_map.keys():
        if data['filial'] == bpm_map[i]:
            data['filial_bpm'] = i
    if data['filial'] == 'АО "Жилищный строительный сберегательный банк "Отбасы банк"':
        data['filial_type'] = 'Пользователи ЦА'
    else:
        data['filial_type'] = 'Пользователи филиалов'

    return data


def replace_code(data):
    codes = {
        'mng': 'Мангистауский ОФ АО «Жилищный строительный сберегательный банк «Отбасы Банк»',
        'akt': 'Актюбинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'alm': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Алматы',
        'ast': 'Центральный филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'atr': 'Атырауский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
        'krg': 'Карагандинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'akm': 'Акмолинский ОФ АО «Жилищный строительный сберегательный банк «Отбасы банк»',
        'kos': 'Костанайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'kzo': 'Кызылординский филиал АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'pvl': 'Павлодарский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'sko': 'Северо-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'sem': 'Абайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'shm': 'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Шымкент',
        'tdk': 'Алматинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"',
        'trz': 'Жамбылский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'trk': 'Туркестанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'zko': 'Западно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"',
        'vko': 'Восточно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"'
    }

    for i in codes.keys():
        if data['filial'] == codes[i]:
            data['filial_c'] = i
    return data


def is_spec_alphabet(username: str):
    if username.split()[1][0].lower() in ['ж', 'ш', 'ч', 'щ']:
        return True
    return False


def change_login(user: dict):
    if 'key' not in user.keys():
        user['key'] = 2
    else:
        user['key'] += 1
    text = translit(user['temp_name'], 'ru', reversed=True).split()
    if is_spec_alphabet(user['username']):
        user['login'] += text[1][user['key']]
        user['user'] += text[1][user['key']]
        user['user'] = user['user'].upper()
        user['mail'] = f'{user["login"]}@otbasybank.kz'  # Почта

    else:
        user['login'] += text[1][user['key'] - 1]
        user['user'] += text[1][user['key'] - 1]
        user['user'] = user['user'].upper()
        user['mail'] = f'{user["login"]}@otbasybank.kz'  # Почта


if __name__ == '__main__':
    ...
